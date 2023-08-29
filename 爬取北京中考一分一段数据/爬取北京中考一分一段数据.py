import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

headers = {
   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

# 请求网页
def get_html(url):
   response = requests.get(url, headers=headers)
   response.encoding = 'utf-8'
   if response.status_code == 200:
       return response.text
   else:
       print(f"Failed to fetch {url}")
       return None

# 解析HTML并提取表格数据
def extract_table_data(html):
   soup = BeautifulSoup(html, "html.parser")
   table = soup.find("table")
   if table is None:
       print("Failed to find the table")
       return None

   data = []
   for row in table.find_all("tr"):
       row_data = [cell.text.strip() for cell in row.find_all("td")]
       data.append(row_data)

   return data

# 提取详细页面链接
def extract_detail_links(html):
   soup = BeautifulSoup(html, "html.parser")
   links = []
   for link in soup.find("table").find_all("a", href=True):
       links.append([link["href"],link.contents[0]])
   return links

# 将表格数据写入Excel文件
def write_to_excel(data,filename,sheet_name):
   try:
       wb = load_workbook(filename)
   except FileNotFoundError:
       wb = Workbook()
   
   ws = wb.create_sheet(sheet_name)
   for row in data:
       ws.append(row)

   wb.save(filename)

# 主函数
def main():
   url = "https://www.bjeea.cn/html/zkzz/tzgg/2020/0727/76251.html"
   outfilename=f"中考一分一段2020.xlsx"
   html = get_html(url)
   if html is None:
       return

   detail_links = extract_detail_links(html)
   data_list = [["城区","分数","本段人数","累计人数"]]

   for link in detail_links:
       detail_html = get_html(f"https://www.bjeea.cn"+link[0])
       if detail_html is None:
           continue

       data = extract_table_data(detail_html)
       if data is None:
           continue

       write_to_excel(data, outfilename ,link[1])
       print(f"{link[1]} data written {outfilename}")

       for x in data:
            x.insert(0,link[1])
       data_list += data[1:]
   write_to_excel(data_list, outfilename,f"数据合并")
   print(f"All data written {outfilename}。")

if __name__ == "__main__":
   main()